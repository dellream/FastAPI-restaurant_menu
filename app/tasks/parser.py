"""
Сервисный слой для парсинга данных из файла xlsx.

Парсер взаимодействует с объектом файла посредством
библиотеки openyxl.
"""
import json

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.config import MENU_FILE_PATH


class ParserRepo:
    """Класс для парсинга объектов из файла xlsx."""

    def __init__(self) -> None:
        self.sheet: Worksheet = load_workbook(filename=MENU_FILE_PATH).active
        self.parse_result: list = []

    def make_dish(self, row: int) -> dict:
        """Собрать словарь с данными о блюде из файла."""
        # Создаем словарь
        dish: dict[str, str | int] = {}
        # Выставляем диапазон рассматриваемых ячеек
        cells: list[Cell] = self.sheet[f'C{row}':f'F{row}'][0]  # type: ignore
        dish['id'] = str(cells[0].value)
        dish['title'] = cells[1].value
        dish['description'] = cells[2].value
        dish['price'] = str(cells[3].value).replace(',', '.')
        # if cells[4].value:
        #     dish['discount'] = cells[4].value
        # else:
        #     dish['discount'] = 0
        return dish

    def make_submenu(self, row: int, max_row: int) -> dict:
        """Собрать словарь с данными о подменю из файла."""
        submenu: dict = {
            'dishes': [],
        }
        # Выставляем диапазон рассматриваемых ячеек
        cells: list[Cell] = self.sheet[f'B{row}':f'D{row}'][0]  # type: ignore
        submenu['id'] = str(cells[0].value)
        submenu['title'] = cells[1].value
        submenu['description'] = cells[2].value
        for i in range(row + 1, max_row + 1):
            if self.sheet[f'C{i}'].value:
                dish = self.make_dish(i)
                if dish['description']:
                    submenu['dishes'].append(dish)
                else:
                    break
        return submenu

    def make_menu(self, row: int, max_row: int) -> dict:
        """Собрать словарь с данными о меню из файла."""
        menu: dict = {
            'submenus': [],
        }
        # Выставляем диапазон рассматриваемых ячеек
        cells: list[Cell] = self.sheet[f'A{row}':f'C{row}'][0]  # type: ignore
        menu['id'] = str(cells[0].value)
        menu['title'] = cells[1].value
        menu['description'] = cells[2].value
        for i in range(row + 1, max_row + 1):
            if self.sheet[f'B{i}'].value:
                submenu = self.make_submenu(i, max_row)
                if submenu['description']:
                    menu['submenus'].append(submenu)
                else:
                    break
        return menu

    def parser(self) -> list[dict[str, str | list]]:
        """Парсинг файла."""
        for i in range(1, self.sheet.max_row + 1):
            if self.sheet[f'A{i}'].value:
                self.parse_result.append(self.make_menu(i, self.sheet.max_row))

        # Добавляем вывод в консоль или в логи
        for inform in self.parse_result:
            print('Parsed item:')
            print(json.dumps(inform, indent=4))  # Вывод в консоль

        return self.parse_result


if __name__ == '__main__':
    parser = ParserRepo()
    parsed_data = parser.parser()
    for item in parsed_data:
        print(item)
